from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path
from threading import Lock
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel

from .models import (
    ACTIVE_STATUSES,
    EmployeeRecord,
    ExpirationTarget,
    RequestRecord,
    STATUS_COMPLETED,
    STATUS_ESCALATED,
    STATUS_IN_PROGRESS,
    STATUS_SUBMITTED,
)


REQUEST_HEADERS = [
    "request_id",
    "title",
    "category",
    "assigned_to",
    "associate_email",
    "manager_email",
    "due_date",
    "status",
    "priority",
    "source_system",
    "requested_change",
    "current_value",
    "submitted_value",
    "notes",
    "created_at",
    "submitted_at",
    "last_reminder_at",
    "reminder_count",
    "escalated_at",
]

EMPLOYEE_HEADERS = [
    "Title",
    "Name",
    "Email",
    "Manager Email",
    "Expirary Date",
]

AUDIT_HEADERS = ["timestamp", "record_id", "actor", "action", "details"]
REMINDER_HEADERS = [
    "timestamp",
    "target_key",
    "employee_id",
    "employee_name",
    "recipient",
    "field_name",
    "expiration_date",
    "subject",
    "channel",
    "status",
    "message_preview",
]
REMINDER_REQUESTS_HEADERS = [
    "timestamp",
    "target_key",
    "employee_id",
    "employee_name",
    "field_name",
    "expiration_date",
    "days_until_due",
    "status",
    "recipient",
    "subject",
    "channel",
]


class ExcelRepository:
    def __init__(self, workbook_path: Path):
        self.workbook_path = Path(workbook_path)
        self._lock = Lock()
        if not self.workbook_path.exists():
            self.create_mock_workbook()
        self.ensure_workbook_ready()

    @staticmethod
    def validate_workbook(workbook_path: Path) -> None:
        workbook = load_workbook(workbook_path)
        if "Employees" not in workbook.sheetnames:
            raise ValueError("Workbook must contain an Employees sheet.")

        headers = [cell.value for cell in workbook["Employees"][1]]
        missing = [header for header in EMPLOYEE_HEADERS if header not in headers]
        if missing:
            raise ValueError(f"Employees sheet is missing columns: {', '.join(missing)}")

    def ensure_workbook_ready(self) -> None:
        self.validate_workbook(self.workbook_path)
        workbook = load_workbook(self.workbook_path)
        changed = False
        if "AuditLog" not in workbook.sheetnames:
            audit = workbook.create_sheet("AuditLog")
            audit.append(AUDIT_HEADERS)
            audit.append([datetime.now(), "SYSTEM", "upload", "created", "Audit sheet initialized."])
            changed = True
        if "ReminderLog" not in workbook.sheetnames:
            workbook.create_sheet("ReminderLog").append(REMINDER_HEADERS)
            changed = True
        if "ReminderRequests" not in workbook.sheetnames:
            workbook.create_sheet("ReminderRequests").append(REMINDER_REQUESTS_HEADERS)
            changed = True
        if changed:
            workbook.save(self.workbook_path)

    def create_mock_workbook(self) -> None:
        self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        employees = workbook.active
        employees.title = "Employees"
        employees.append(EMPLOYEE_HEADERS)

        today = date.today()
        rows = [
            ["Gmail SMTP Test Reminder", "Chinmay Dave", "chinmaydavecs@gmail.com", "chinmaydaveatl@gmail.com", today],
            ["Driver License Renewal", "Blake Smith", "blake.smith@example.com", "maria.bennett@example.com", today - timedelta(days=7)],
            ["Work Visa Renewal", "Casey Lee", "casey.lee@example.com", "maria.bennett@example.com", today - timedelta(days=2)],
            ["Professional License Renewal", "Devon Patel", "devon.patel@example.com", "sarah.nguyen@example.com", today],
            ["Passport Renewal", "Emerson Clark", "emerson.clark@example.com", "sarah.nguyen@example.com", today + timedelta(days=1)],
            ["Commercial Driver License Renewal", "Finley Morgan", "finley.morgan@example.com", "sarah.nguyen@example.com", today + timedelta(days=2)],
            ["Security Badge Expiration", "Gray Taylor", "gray.taylor@example.com", "james.carter@example.com", today + timedelta(days=3)],
            ["Medical License Renewal", "Harper Davis", "harper.davis@example.com", "james.carter@example.com", today + timedelta(days=5)],
            ["Passport Renewal", "Indigo Martinez", "indigo.martinez@example.com", "james.carter@example.com", today + timedelta(days=7)],
            ["Work Permit Renewal", "Jordan Brown", "jordan.brown@example.com", "olivia.reed@example.com", today + timedelta(days=10)],
            ["Driver License Renewal", "Kai Wilson", "kai.wilson@example.com", "olivia.reed@example.com", today + timedelta(days=14)],
            ["Passport Renewal", "Logan Anderson", "logan.anderson@example.com", "olivia.reed@example.com", today + timedelta(days=18)],
            ["Professional Certification Renewal", "Morgan Thomas", "morgan.thomas@example.com", "ethan.brooks@example.com", today + timedelta(days=21)],
            ["Visa Document Review", "Noa White", "noa.white@example.com", "ethan.brooks@example.com", today + timedelta(days=25)],
            ["Driver License Renewal", "Oakley Harris", "oakley.harris@example.com", "ethan.brooks@example.com", today + timedelta(days=30)],
            ["Passport Renewal", "Parker Martin", "parker.martin@example.com", "nina.powers@example.com", today + timedelta(days=35)],
            ["State License Renewal", "Quinn Thompson", "quinn.thompson@example.com", "nina.powers@example.com", today + timedelta(days=45)],
            ["Work Visa Renewal", "Riley Garcia", "riley.garcia@example.com", "nina.powers@example.com", today + timedelta(days=60)],
            ["Passport Renewal", "Sawyer Robinson", "sawyer.robinson@example.com", "marcus.hill@example.com", today + timedelta(days=75)],
            ["Driver License Renewal", "Taylor Lewis", "taylor.lewis@example.com", "marcus.hill@example.com", today + timedelta(days=90)],
            ["Professional License Renewal", "Uma Walker", "uma.walker@example.com", "marcus.hill@example.com", today + timedelta(days=120)],
            ["Passport Renewal", "Val Jordan", "val.jordan@example.com", "rachel.kim@example.com", today + timedelta(days=150)],
            ["Security Clearance Renewal", "Winter Hall", "winter.hall@example.com", "rachel.kim@example.com", today + timedelta(days=180)],
            ["Driver License Renewal", "Xavier Young", "xavier.young@example.com", "rachel.kim@example.com", today + timedelta(days=240)],
            ["Passport Renewal", "Yael Allen", "yael.allen@example.com", "rachel.kim@example.com", today + timedelta(days=365)],
        ]

        for row in rows:
            employees.append(row)

        audit = workbook.create_sheet("AuditLog")
        audit.append(AUDIT_HEADERS)
        audit.append([datetime.now(), "SYSTEM", "seed", "created", "Mock workbook initialized."])

        workbook.create_sheet("ReminderLog").append(REMINDER_HEADERS)
        workbook.create_sheet("ReminderRequests").append(REMINDER_REQUESTS_HEADERS)

        workbook.save(self.workbook_path)

    def list_employees(self) -> list[EmployeeRecord]:
        with self._lock:
            workbook = load_workbook(self.workbook_path)
            sheet = workbook["Employees"]
            return [self._row_to_employee(row) for row in self._iter_rows(sheet)]

    def list_requests(self) -> list[RequestRecord]:
        with self._lock:
            workbook = load_workbook(self.workbook_path)
            if "Requests" not in workbook.sheetnames:
                return []
            sheet = workbook["Requests"]
            return [self._row_to_record(row) for row in self._iter_rows(sheet)]

    def get_request(self, request_id: str) -> RequestRecord | None:
        return next((record for record in self.list_requests() if record.request_id == request_id), None)

    def pending_for_reminder(self, days_ahead: int, expiration_filter: str | None = None) -> list[ExpirationTarget]:
        today = date.today()
        if expiration_filter == "overdue":
            def include_target(target: ExpirationTarget) -> bool:
                return target.expiration_date < today
        elif expiration_filter == "7":
            def include_target(target: ExpirationTarget) -> bool:
                return 0 <= target.days_until_due <= 7
        elif expiration_filter == "14":
            def include_target(target: ExpirationTarget) -> bool:
                return 0 <= target.days_until_due <= 14
        elif expiration_filter == "30":
            def include_target(target: ExpirationTarget) -> bool:
                return 0 <= target.days_until_due <= 30
        else:
            horizon = today + timedelta(days=days_ahead)
            def include_target(target: ExpirationTarget) -> bool:
                return target.expiration_date <= horizon

        return [
            target
            for employee in self.list_employees()
            for target in employee.expiration_targets()
            if include_target(target)
        ]

    def find_target_by_key(self, target_key: str) -> ExpirationTarget | None:
        for employee in self.list_employees():
            for target in employee.expiration_targets():
                if target.target_key == target_key:
                    return target
        return None

    def last_reminder_for_target(self, target_key: str) -> datetime | None:
        with self._lock:
            workbook = load_workbook(self.workbook_path)
            if "ReminderLog" not in workbook.sheetnames:
                return None
            sheet = workbook["ReminderLog"]
            headers = [cell.value for cell in sheet[1]]
            rows = self._iter_rows(sheet)
            timestamps = [
                row["timestamp"]
                for row in rows
                if row.get("target_key") == target_key and row.get("timestamp") is not None
            ]
            if not timestamps:
                return None
            return max(timestamps)

    def submit_response(self, request_id: str, submitted_value: str, notes: str, actor: str) -> None:
        with self._lock:
            workbook = load_workbook(self.workbook_path)
            sheet = workbook["Requests"]
            row_idx = self._find_row(sheet, request_id)
            if row_idx is None:
                raise KeyError(f"Request {request_id} not found")

            self._set(sheet, row_idx, "submitted_value", submitted_value)
            self._set(sheet, row_idx, "notes", notes)
            self._set(sheet, row_idx, "status", STATUS_SUBMITTED)
            self._set(sheet, row_idx, "submitted_at", datetime.now())
            self._append_audit(workbook, request_id, actor, "submitted", "Associate submitted form response.")
            workbook.save(self.workbook_path)

    def update_status(self, request_id: str, status: str, actor: str, details: str = "") -> None:
        valid_statuses = {STATUS_IN_PROGRESS, STATUS_SUBMITTED, STATUS_COMPLETED, STATUS_ESCALATED}
        if status not in valid_statuses:
            raise ValueError(f"Unsupported status: {status}")

        with self._lock:
            workbook = load_workbook(self.workbook_path)
            sheet = workbook["Requests"]
            row_idx = self._find_row(sheet, request_id)
            if row_idx is None:
                raise KeyError(f"Request {request_id} not found")
            self._set(sheet, row_idx, "status", status)
            if status == STATUS_ESCALATED:
                self._set(sheet, row_idx, "escalated_at", datetime.now())
            self._append_audit(workbook, request_id, actor, "status_updated", details or f"Status changed to {status}.")
            workbook.save(self.workbook_path)

    def record_reminder(
        self,
        target: ExpirationTarget,
        recipient: str,
        subject: str,
        message_preview: str,
        channel: str,
        status: str,
    ) -> None:
        with self._lock:
            workbook = load_workbook(self.workbook_path)
            reminders = workbook["ReminderLog"]
            reminders.append([
                datetime.now(),
                target.target_key,
                target.employee_id,
                target.employee_name,
                recipient,
                target.field_name,
                target.expiration_date,
                subject,
                channel,
                status,
                message_preview[:180],
            ])
            self._append_audit(
                workbook,
                target.employee_id,
                "system",
                "reminder_sent",
                f"{channel} reminder recorded for {recipient} ({target.field_name}).",
            )
            workbook.save(self.workbook_path)

    def record_reminder_request(self, target: ExpirationTarget, recipient: str, subject: str, channel: str) -> None:
        with self._lock:
            workbook = load_workbook(self.workbook_path)
            requests = workbook["ReminderRequests"]
            requests.append([
                datetime.now(),
                target.target_key,
                target.employee_id,
                target.employee_name,
                target.field_name,
                target.expiration_date,
                target.days_until_due,
                target.status,
                recipient,
                subject,
                channel,
            ])
            self._append_audit(
                workbook,
                target.employee_id,
                "system",
                "reminder_requested",
                f"Reminder request recorded for {target.field_name}.",
            )
            workbook.save(self.workbook_path)

    def audit_rows(self) -> list[dict[str, Any]]:
        return self._sheet_rows("AuditLog")

    def reminder_rows(self) -> list[dict[str, Any]]:
        return self._sheet_rows("ReminderLog")

    def summary(self) -> dict[str, int]:
        records = self.list_employees()
        expiration_targets = [
            target
            for employee in records
            for target in employee.expiration_targets()
        ]
        return {
            "total": len(records),
            "overdue": sum(1 for target in expiration_targets if target.is_overdue),
            "due_soon": sum(1 for target in expiration_targets if not target.is_overdue and target.days_until_due <= 3),
            "upcoming": sum(1 for target in expiration_targets if target.days_until_due > 3),
        }

    def _sheet_rows(self, sheet_name: str) -> list[dict[str, Any]]:
        with self._lock:
            workbook = load_workbook(self.workbook_path)
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in sheet[1]]
            rows = []
            for raw_row in sheet.iter_rows(min_row=2, values_only=True):
                if any(value is not None for value in raw_row):
                    rows.append(dict(zip(headers, raw_row)))
            return rows

    def _iter_rows(self, sheet: Any) -> list[dict[str, Any]]:
        headers = [cell.value for cell in sheet[1]]
        return [
            dict(zip(headers, raw_row))
            for raw_row in sheet.iter_rows(min_row=2, values_only=True)
            if any(value is not None for value in raw_row)
        ]

    def _row_to_employee(self, row: dict[str, Any]) -> EmployeeRecord:
        def normalize_date(value: Any) -> date | None:
            if value in (None, ""):
                return None
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            if isinstance(value, (int, float)):
                return from_excel(value).date()
            if isinstance(value, str):
                for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
                    try:
                        return datetime.strptime(value.strip(), fmt).date()
                    except ValueError:
                        pass
                raise ValueError(f"Unsupported date format: {value}")
            return value

        return EmployeeRecord(
            title=row["Title"],
            name=row["Name"],
            email=row["Email"],
            manager_email=row["Manager Email"],
            expiration_date=normalize_date(row["Expirary Date"]),
        )

    def _row_to_record(self, row: dict[str, Any]) -> RequestRecord:
        due_date = row["due_date"]
        if isinstance(due_date, datetime):
            due_date = due_date.date()
        return RequestRecord(
            request_id=row["request_id"],
            title=row["title"],
            category=row["category"],
            assigned_to=row["assigned_to"],
            associate_email=row["associate_email"],
            manager_email=row["manager_email"],
            due_date=due_date,
            status=row["status"],
            priority=row["priority"],
            source_system=row["source_system"],
            requested_change=row["requested_change"],
            current_value=row["current_value"] or "",
            submitted_value=row["submitted_value"] or "",
            notes=row["notes"] or "",
            created_at=row["created_at"],
            submitted_at=row["submitted_at"],
            last_reminder_at=row["last_reminder_at"],
            reminder_count=int(row["reminder_count"] or 0),
            escalated_at=row["escalated_at"],
        )

    def _find_row(self, sheet: Any, request_id: str) -> int | None:
        request_col = REQUEST_HEADERS.index("request_id") + 1
        for row_idx in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_idx, column=request_col).value == request_id:
                return row_idx
        return None

    def _set(self, sheet: Any, row_idx: int, field: str, value: Any) -> None:
        sheet.cell(row=row_idx, column=REQUEST_HEADERS.index(field) + 1, value=value)

    def _get(self, sheet: Any, row_idx: int, field: str) -> Any:
        return sheet.cell(row=row_idx, column=REQUEST_HEADERS.index(field) + 1).value

    def _append_audit(self, workbook: Any, request_id: str, actor: str, action: str, details: str) -> None:
        workbook["AuditLog"].append([datetime.now(), request_id, actor, action, details])
